from flask import Blueprint, jsonify, request, send_file
from bson import ObjectId
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
from database import db

quote_routes = Blueprint('quote_routes', __name__)

@quote_routes.route('/api/quotes/<quote_id>', methods=['PUT'])
def update_quote(quote_id):
    try:
        data = request.get_json()
        
        # Convert string date to datetime object
        if 'date' in data and data['date']:
            try:
                data['date'] = datetime.strptime(data['date'], '%Y-%m-%d')
            except ValueError:
                pass  # If date conversion fails, keep the original value
        
        # Update the document
        result = db.quotes.update_one(
            {'_id': ObjectId(quote_id)},
            {'$set': data}
        )
        
        if result.modified_count > 0:
            return jsonify({'success': True, 'message': 'Quote updated successfully'})
        else:
            return jsonify({'success': False, 'message': 'No changes made to the quote'}), 400
            
    except Exception as e:
        print("Error updating quote:", str(e))
        return jsonify({'error': str(e)}), 500

@quote_routes.route('/api/quotes/<quote_id>', methods=['DELETE'])
def delete_quote(quote_id):
    try:
        result = db.quotes.delete_one({'_id': ObjectId(quote_id)})
        return jsonify({'success': True if result.deleted_count > 0 else False})
    except Exception as e:
        print("Error deleting quote:", str(e))
        return jsonify({'error': str(e)}), 500

@quote_routes.route('/create_quote', methods=['POST'])
def create_quote():
    try:
        data = request.get_json()
        
        # Convert date string to datetime object
        if 'date' in data and data['date']:
            try:
                data['date'] = datetime.strptime(data['date'], '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': 'Invalid date format'}), 400
        
        # Insert the new quote
        result = db.quotes.insert_one(data)
        
        if result.inserted_id:
            return jsonify({
                'success': True,
                'message': 'Quote created successfully',
                'quote_id': str(result.inserted_id)
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to create quote'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@quote_routes.route('/api/leads/<lead_id>', methods=['DELETE'])
def delete_lead(lead_id):
    try:
        result = db.leads.delete_one({'_id': ObjectId(lead_id)})
        return jsonify({'success': True if result.deleted_count > 0 else False})
    except Exception as e:
        print("Error deleting lead:", str(e))
        return jsonify({'error': str(e)}), 500

@quote_routes.route('/export-quotes')
def export_quotes():
    # Get all quotes from the database and print first one for debugging
    quotes = list(db.quotes.find().sort('Date', 1))
    if quotes:
        print("Debug - Sample quote data:", quotes[0])
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes Report"
    
    # Define headers and their corresponding database fields
    headers_mapping = [
        ('Reference', 'Ref'),
        ('Date', 'Date'),
        ('Consultant', 'Cons'),
        ('Client', 'Client'),
        ('Source', 'Source'),
        ('Suburb', 'Suburb'),
        ('Quote Value', 'Quote Value'),
        ('Sold Value', 'Sold Value'),
        ('Comments', 'Comments')
    ]
    
    # Write headers and apply styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
    
    for col, (header, _) in enumerate(headers_mapping, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row, quote in enumerate(quotes, 2):
        # Debug print for each quote
        print(f"Processing quote {row-1}:", quote)
        
        for col, (_, field) in enumerate(headers_mapping, 1):
            cell = ws.cell(row=row, column=col)
            value = quote.get(field, '')
            
            # Format specific fields
            if field == 'Date':
                if value:
                    try:
                        # Parse the date string
                        date_obj = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                        cell.value = date_obj.date()
                        cell.number_format = 'YYYY-MM-DD'
                    except (ValueError, TypeError):
                        cell.value = value
            elif field in ['Quote Value', 'Sold Value']:
                if value:
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        cell.value = 0.0
                else:
                    cell.value = 0.0
            else:
                cell.value = str(value) if value is not None else ''
    
    # Auto-adjust column widths
    for col in range(1, len(headers_mapping) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        column = ws[column_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Create 'exports' directory if it doesn't exist
    if not os.path.exists('exports'):
        os.makedirs('exports')
    
    # Save the file with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'exports/quotes_report_{timestamp}.xlsx'
    wb.save(filename)
    
    # Send file
    return send_file(
        filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'quotes_report_{timestamp}.xlsx'
    )
